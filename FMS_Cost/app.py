from io import BytesIO

import openpyxl
from flask import Flask, render_template, request, jsonify, send_file

import config
import reference_data
import matching
from matching import strip_platform_suffix
import learned_mappings
import report_generator

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = config.MAX_CONTENT_LENGTH

# Load reference data at startup
reference_data.load_references()


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Файл не загружен"}), 400

    date_str = request.form.get("date", "")
    exchange_rate = request.form.get("exchange_rate", "")
    if not date_str or not exchange_rate:
        return jsonify({"error": "Укажите дату и курс доллара"}), 400

    try:
        exchange_rate = float(exchange_rate)
    except ValueError:
        return jsonify({"error": "Курс доллара должен быть числом"}), 400

    if exchange_rate <= 0:
        return jsonify({"error": "Курс доллара должен быть больше нуля"}), 400

    # Parse uploaded Excel file
    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
    except Exception:
        return jsonify({"error": "Не удалось прочитать файл. Убедитесь, что это .xlsx"}), 400

    ws = wb.active

    # Find columns by header names (scan first 5 rows)
    offer_col = None
    source_col = None
    cost_col = None
    header_row = None

    for row_idx in range(1, min(6, ws.max_row + 1)):
        for cell in ws[row_idx]:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip()
                if val == config.STATS_OFFER_HEADER:
                    offer_col = cell.column - 1
                    header_row = row_idx
                elif val == config.STATS_SOURCE_HEADER:
                    source_col = cell.column - 1
                elif val == config.STATS_COST_HEADER:
                    cost_col = cell.column - 1

    if offer_col is None or source_col is None or cost_col is None:
        wb.close()
        return jsonify({
            "error": "Не найдены обязательные столбцы: Оффер, Источник, Расход"
        }), 400

    # Parse data rows
    stats_rows = []
    for row in ws.iter_rows(
        min_row=header_row + 1, max_row=ws.max_row, values_only=True
    ):
        offer_name = row[offer_col] if offer_col < len(row) else None
        source_name = row[source_col] if source_col < len(row) else None
        cost_rub = row[cost_col] if cost_col < len(row) else None

        if offer_name and source_name:
            stats_rows.append({
                "offer_name": str(offer_name).strip(),
                "source_name": str(source_name).strip(),
                "cost_rub": float(cost_rub) if cost_rub else 0.0,
            })

    wb.close()

    if not stats_rows:
        return jsonify({"error": "Файл не содержит данных"}), 400

    # Attempt matching
    offers_dict = reference_data.get_offers()
    sources_dict = reference_data.get_sources()

    unique_offers = set(r["offer_name"] for r in stats_rows)
    unique_sources = set(r["source_name"] for r in stats_rows)

    matched_offers = {}
    unmatched_offers = []
    for name in sorted(unique_offers):
        offer_id = matching.match_offer(name, offers_dict)
        if offer_id is not None:
            matched_offers[name] = offer_id
        else:
            unmatched_offers.append(name)

    matched_sources = {}
    unmatched_sources = []
    for name in sorted(unique_sources):
        source_id = matching.match_source(name, sources_dict)
        if source_id is not None:
            matched_sources[name] = source_id
        else:
            unmatched_sources.append(name)

    return jsonify({
        "success": True,
        "stats_rows": stats_rows,
        "matched_offers": matched_offers,
        "matched_sources": matched_sources,
        "unmatched_offers": unmatched_offers,
        "unmatched_sources": unmatched_sources,
    })


@app.route("/save-mappings", methods=["POST"])
def save_mappings():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Нет данных"}), 400

    manual_offers = data.get("offers", {})
    manual_sources = data.get("sources", {})
    if manual_offers:
        learned_mappings.save_offers(manual_offers)
        for name, offer_id in manual_offers.items():
            reference_data.add_offer(int(offer_id), strip_platform_suffix(name))
    if manual_sources:
        learned_mappings.save_sources(manual_sources)
        for name, source_id in manual_sources.items():
            reference_data.add_source(int(source_id), name)

    return jsonify({"success": True})


@app.route("/generate", methods=["POST"])
def generate():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Нет данных"}), 400

    date_str = data.get("date", "")
    exchange_rate = float(data.get("exchange_rate", 0))
    stats_rows = data.get("stats_rows", [])
    offer_ids = data.get("offer_ids", {})
    source_ids = data.get("source_ids", {})

    if not stats_rows or not exchange_rate:
        return jsonify({"error": "Недостаточно данных для генерации"}), 400

    # Save manually entered IDs for future use
    manual_offers = data.get("manual_offer_ids", {})
    manual_sources = data.get("manual_source_ids", {})
    if manual_offers:
        learned_mappings.save_offers(manual_offers)
        for name, offer_id in manual_offers.items():
            reference_data.add_offer(int(offer_id), strip_platform_suffix(name))
    if manual_sources:
        learned_mappings.save_sources(manual_sources)
        for name, source_id in manual_sources.items():
            reference_data.add_source(int(source_id), name)

    report_bytes = report_generator.generate_report(
        date_str=date_str,
        exchange_rate=exchange_rate,
        stats_rows=stats_rows,
        offer_id_map=offer_ids,
        source_id_map=source_ids,
        no_convert_sources=config.NO_CONVERT_SOURCES,
    )

    return send_file(
        BytesIO(report_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"report_{date_str}.xlsx",
    )


# --- Reference pages ---

@app.route("/offers")
def offers_page():
    return render_template("reference.html", type="offers", title="Оферы",
                           name_label="Название офера", id_label="ID офера")


@app.route("/sources")
def sources_page():
    return render_template("reference.html", type="sources", title="Источники",
                           name_label="Название источника", id_label="ID источника")


@app.route("/api/offers")
def api_offers():
    return jsonify(reference_data.get_offers_list())


@app.route("/api/sources")
def api_sources():
    return jsonify(reference_data.get_sources_list())


@app.route("/api/offers", methods=["POST"])
def api_add_offer():
    data = request.get_json()
    try:
        offer_id = int(data["id"])
        name = data["name"].strip()
    except (KeyError, ValueError, TypeError):
        return jsonify({"error": "Укажите ID (число) и название"}), 400
    if not name:
        return jsonify({"error": "Название не может быть пустым"}), 400
    reference_data.add_offer(offer_id, name)
    return jsonify({"success": True})


@app.route("/api/offers/<int:offer_id>", methods=["DELETE"])
def api_delete_offer(offer_id):
    reference_data.delete_offer(offer_id)
    return jsonify({"success": True})


@app.route("/api/sources", methods=["POST"])
def api_add_source():
    data = request.get_json()
    try:
        source_id = int(data["id"])
        name = data["name"].strip()
    except (KeyError, ValueError, TypeError):
        return jsonify({"error": "Укажите ID (число) и название"}), 400
    if not name:
        return jsonify({"error": "Название не может быть пустым"}), 400
    reference_data.add_source(source_id, name)
    return jsonify({"success": True})


@app.route("/api/sources/<int:source_id>", methods=["DELETE"])
def api_delete_source(source_id):
    reference_data.delete_source(source_id)
    return jsonify({"success": True})


if __name__ == "__main__":
    app.run(debug=True, port=5001)

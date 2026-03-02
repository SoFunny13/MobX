import openpyxl
import config

# Module-level caches: normalized_name -> (id, original_name)
_offers = {}
_sources = {}


def load_references():
    """Load both reference files into module-level caches."""
    global _offers, _sources
    _offers = _load_offers()
    _sources = _load_sources()


def _load_offers():
    wb = openpyxl.load_workbook(config.OFFERS_FILE, read_only=True, data_only=True)
    ws = wb[config.OFFERS_SHEET]
    result = {}
    for row in ws.iter_rows(
        min_row=config.OFFERS_DATA_START_ROW,
        max_row=ws.max_row,
        values_only=True,
    ):
        offer_id = row[config.OFFERS_COL_ID]
        name = row[config.OFFERS_COL_NAME]
        if offer_id is not None and name:
            normalized = _normalize(str(name))
            result[normalized] = (int(offer_id), str(name))
    wb.close()
    return result


def _load_sources():
    wb = openpyxl.load_workbook(config.SOURCES_FILE, read_only=True, data_only=True)
    ws = wb[config.SOURCES_SHEET]
    result = {}
    for row in ws.iter_rows(
        min_row=config.SOURCES_DATA_START_ROW,
        max_row=ws.max_row,
        values_only=True,
    ):
        source_id = row[config.SOURCES_COL_ID]
        name = row[config.SOURCES_COL_NAME]
        if source_id is not None and name:
            normalized = _normalize(str(name))
            result[normalized] = (int(source_id), str(name))
    wb.close()
    return result


def _normalize(name):
    return " ".join(name.lower().strip().split())


def get_offers():
    """Return dict: normalized_name -> (id, original_name)."""
    return _offers


def get_sources():
    """Return dict: normalized_name -> (id, original_name)."""
    return _sources


def get_offers_list():
    """Return sorted list of dicts [{id, name}, ...]."""
    return sorted(
        [{"id": oid, "name": name} for _, (oid, name) in _offers.items()],
        key=lambda x: x["name"],
    )


def get_sources_list():
    """Return sorted list of dicts [{id, name}, ...]."""
    return sorted(
        [{"id": sid, "name": name} for _, (sid, name) in _sources.items()],
        key=lambda x: x["name"],
    )


def add_offer(offer_id, name):
    wb = openpyxl.load_workbook(config.OFFERS_FILE)
    ws = wb[config.OFFERS_SHEET]
    ws.append([offer_id, name])
    wb.save(config.OFFERS_FILE)
    wb.close()
    global _offers
    _offers = _load_offers()


def delete_offer(offer_id):
    wb = openpyxl.load_workbook(config.OFFERS_FILE)
    ws = wb[config.OFFERS_SHEET]
    for row_idx in range(ws.max_row, 0, -1):
        cell_val = ws.cell(row=row_idx, column=config.OFFERS_COL_ID + 1).value
        if cell_val is not None and int(cell_val) == offer_id:
            ws.delete_rows(row_idx)
            break
    wb.save(config.OFFERS_FILE)
    wb.close()
    global _offers
    _offers = _load_offers()


def add_source(source_id, name):
    wb = openpyxl.load_workbook(config.SOURCES_FILE)
    ws = wb[config.SOURCES_SHEET]
    ws.append([source_id, name])
    wb.save(config.SOURCES_FILE)
    wb.close()
    global _sources
    _sources = _load_sources()


def delete_source(source_id):
    wb = openpyxl.load_workbook(config.SOURCES_FILE)
    ws = wb[config.SOURCES_SHEET]
    for row_idx in range(ws.max_row, 0, -1):
        cell_val = ws.cell(row=row_idx, column=config.SOURCES_COL_ID + 1).value
        if cell_val is not None and int(cell_val) == source_id:
            ws.delete_rows(row_idx)
            break
    wb.save(config.SOURCES_FILE)
    wb.close()
    global _sources
    _sources = _load_sources()
